import openpyxl
import json


def load_data():
    weekly_data = {}
    monthly_data = {}
    annual_data = {}
    data = [weekly_data, monthly_data, annual_data]

    wb = openpyxl.load_workbook(r"T:\Research & Project\Historical Rate of Return.xlsm", data_only=True)
    #ws_asset_mix = wb['Asset Mix']
    #ws_dividend_scale = wb['Dividend Scale']
    ws_market_rates = wb['Market Rates']
    #ws_gia = wb['Guaranteed Interest Account']

    ### Market Rates ###
    max_row = 0
    market_rates_values = [[]]
    # Iterate over columns B to M
    for col in range(2, 14):  # 2 corresponds to column B and 13 to column M
        column = ws_market_rates[ws_market_rates.cell(row=1, column=col).column_letter]
        max_row_in_column = max([cell.row for cell in column if cell.value is not None])
        max_row = max(max_row, max_row_in_column)
        market_rates_values.append([])
    for i in range(4, max_row + 1):
        for col in range(2, 14):
            value = ws_market_rates.cell(i, col).value * 100 if ws_market_rates.cell(i, col).value is not None else None
            market_rates_values[col - 2].append(value)
    weekly_data = {
        'ba_1month': market_rates_values[0],
        'ba_3month': market_rates_values[1],
        'prime': market_rates_values[2],
        'corra': market_rates_values[3],
        'ca2y': market_rates_values[4],
        'ca5y': market_rates_values[5],
        'ca10y': market_rates_values[6],
        'ca30y': market_rates_values[7],
        'ca10y_2y_spread': market_rates_values[8],
        'gic_5y': market_rates_values[9],
        'cpi_headline': market_rates_values[10],
        'cpi_core': market_rates_values[11]
    }

    with open('C:\Users\wqiao\OneDrive - Pal Insurance Services LTD\Desktop\Useful Materials\data.json', 'w', encoding='utf-8') as f:
        json.dump(weekly_data, f, ensure_ascii=False, indent=4)


if __name__ == "__main__":
    load_data()
